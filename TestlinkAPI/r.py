from openpyxl import load_workbook
from testlink import TestlinkAPIClient
import re
from Utils.Inputs.Common_inputs import inputs


class Test:
    tlc = TestlinkAPIClient(inputs.API_URL, inputs.KEY)

    def read_test_case_excel(self, file_path):
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = {cell.value.strip(): idx for idx, cell in enumerate(sheet[1])}

        required_columns = [
            "Category", "Test case ID", "Description", "Keywords",
            "Steps_actions", "expected_results", "execution_type", "Test cases"
        ]
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")

        test_case_data = []
        current_test_case = None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            category, test_case_id, description, keywords, steps_actions, expected_results, execution_type, summary = \
                [row[headers[col]] for col in required_columns]

            # Handle new test case
            if test_case_id:
                if current_test_case:
                    test_case_data.append(current_test_case)

                current_test_case = {
                    "Category": category, "Test Case ID": test_case_id, "Description": description,
                    "Keywords": keywords, "Expected Output": expected_results, "Steps Data": [], "Summary": summary
                }

            # Process steps actions
            if steps_actions:
                steps = [re.sub(r'^\d+\.\s*', '', step.strip()) for step in steps_actions.split("\n") if step.strip()]
                for step in steps:
                    current_test_case["Steps Data"].append({
                        "step_number": len(current_test_case["Steps Data"]) + 1,
                        "actions": step, "expected_results": expected_results, "execution_type": execution_type
                    })

        if current_test_case:
            test_case_data.append(current_test_case)

        for test_case in test_case_data:
            self.upload_test_case_to_testlink(test_case)

    def upload_test_case_to_testlink(self, test_case):
        suite_id = self.get_or_create_test_suite(test_case['Category'])
        project_id = self.get_project_id('PCI')

        case_name = f"{test_case['Test Case ID']} - {test_case['Description']}"
        if self.test_case_exists(case_name, suite_id):
            print(f"Test case '{case_name}' exists. Skipping.")
            return

        steps_list = test_case['Steps Data']
        test_case_response = self.tlc.createTestCase(
            testcasename=case_name, testsuiteid=suite_id, testprojectid=project_id,
            authorlogin="admin", summary=test_case['Summary'], steps=steps_list,
            expectedresults=test_case['Expected Output']
        )

        test_case_id = test_case_response[0]['id']
        tc_full_ext_id = self.tlc.getTestCase(testcaseid=test_case_id)[0]["full_tc_external_id"]
        keywords_list = [kw.strip() for kw in test_case['Keywords'].split(",") if kw.strip()]
        if keywords_list:
            self.tlc.addTestCaseKeywords({tc_full_ext_id: keywords_list})
        print(f"Test case '{case_name}' created and keywords added.")

    def test_case_exists(self, test_case_name, suite_id):
        return any(test_case["name"].strip().lower() == test_case_name.strip().lower()
                   for test_case in self.tlc.getTestCasesForTestSuite(suite_id, False, False))

    def get_project_id(self, project_name):
        project = next((proj for proj in self.tlc.getProjects() if proj["name"] == project_name), None)
        if not project:
            raise ValueError(f"Project '{project_name}' not found.")
        return project["id"]

    def get_or_create_test_suite(self, suite_name):
        project_id = self.get_project_id("PCI")
        suite = next((suite for suite in self.tlc.getFirstLevelTestSuitesForTestProject(project_id)
                      if suite["name"] == suite_name), None)
        if suite:
            return suite["id"]
        suite = self.tlc.createTestSuite(project_id, suite_name, "Imported test suite")
        return suite[0]["id"]


if __name__ == '__main__':
    ts = Test()
    ts.read_test_case_excel(inputs.EXCEL_PATH)
