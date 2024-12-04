from openpyxl import load_workbook
from testlink import TestlinkAPIClient
import re
from Utils.Inputs.Common_inputs import inputs


class Test:
    tlc = TestlinkAPIClient(inputs.API_URL, inputs.KEY)

    def read_test_case_excel(self, file_path):
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active
        headers = {cell.value.strip(): idx for idx, cell in enumerate(sheet[1])}

        required_columns = [
            "Category", "Test case ID", "Description", "Keywords",
            "Steps_actions", "expected_results", "execution_type", "Test cases"
        ]
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"Missing required column: {col}")


        test_case_steps=[]
        current_step_number = 1



        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Directly unpack values from the row
            category, test_case_id, description, keywords, steps_actions, expected_results, execution_type, summary = \
                [row[headers[col]] for col in required_columns]

            if test_case_id:  # Process only rows that contain a test case ID
                self.upload_test_case_to_testlink(category, test_case_id, description, keywords,
                                                  steps_actions, expected_results, execution_type, summary)

            current_step_number = 1
            step_list = []
            # Collect the steps for this test case
            if steps_actions:
                steps = [re.sub(r'^\d+\.\s*', '', step.strip()) for step in steps_actions.split("\n") if step.strip()]

                for action in steps:
                    step = {
                        "step_number": current_step_number,
                        "actions": action,
                        "expected_results": expected_results,
                        "execution_type": execution_type
                    }
                    step_list.append(step)
                    current_step_number += 1
            if step_list:
                test_case_steps.append(step_list)




    def upload_test_case_to_testlink(self, category, test_case_id, description, keywords, steps_actions,
                                     expected_results, execution_type, summary):
        suite_id = self.get_or_create_test_suite(category)
        project_id = self.get_project_id('PCI')

        case_name = f"{test_case_id} - {description}"
        if self.test_case_exists(case_name, suite_id):
            print(f"Test case '{case_name}' exists. Skipping.")
            return

        # Create the test case in TestLink
        test_case_response = self.tlc.createTestCase(
            testcasename=case_name,
            testsuiteid=suite_id,
            testprojectid=project_id,
            authorlogin="admin",
            summary=summary,
            steps=test_case_steps,
            expectedresults=expected_results
        )

        test_case_id = test_case_response[0]['id']
        tc_full_ext_id = self.tlc.getTestCase(testcaseid=test_case_id)[0]["full_tc_external_id"]
        keywords_list = [kw.strip() for kw in keywords.split(",") if kw.strip()]
        if keywords_list:
            self.tlc.addTestCaseKeywords({tc_full_ext_id: keywords_list})
        print(f"Test case '{case_name}' created and keywords added.")

    def test_case_exists(self, test_case_name, suite_id):
        return any(test_case["name"].strip().lower() == test_case_name.strip().lower()
                   for test_case in self.tlc.getTestCasesForTestSuite(suite_id, False, False))

    def get_project_id(self, project_name):
        project = next((proj for proj in self.tlc.getProjects() if proj["name"] == project_name), None)
        if not project:
            raise ValueError(f"Project '{project_name}' not found.")
        return project["id"]

    def get_or_create_test_suite(self, suite_name):
        project_id = self.get_project_id("PCI")
        suite = next((suite for suite in self.tlc.getFirstLevelTestSuitesForTestProject(project_id)
                      if suite["name"] == suite_name), None)
        if suite:
            return suite["id"]
        suite = self.tlc.createTestSuite(project_id, suite_name, "Imported test suite")
        return suite[0]["id"]


if __name__ == '__main__':
    ts = Test()
    ts.read_test_case_excel(inputs.EXCEL_PATH)
