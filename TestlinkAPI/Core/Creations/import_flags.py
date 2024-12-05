from openpyxl import load_workbook
from testlink import TestlinkAPIClient
import re
from Utils.Inputs.Common_inputs import inputs



class Test:
    tlc = TestlinkAPIClient(inputs.API_URL, inputs.KEY)


    def read_test_case_excel(self,file_path):

        workbook = load_workbook(filename=file_path)
        sheet = workbook.active


        headers = {cell.value.strip(): idx for idx, cell in enumerate(sheet[1])}
        required_columns = [
            "Category", "Test case ID", "Description", "preconditions" ,"Keywords",
            "status", "importance" , "exec time" , "Steps_actions", "expected_results", "execution_type", "Test cases"
        ]

        for column in required_columns:
            if column not in headers:
                raise ValueError(f"Missing required column: {column}")

        test_case_data = []
        current_test_case = {}
        current_step_number = 1


        for row in sheet.iter_rows(min_row=2, values_only=True):
            category = row[headers["Category"]]
            test_case_id = row[headers["Test case ID"]]
            description = row[headers["Description"]]
            preconditions = row[headers["preconditions"]]
            keywords = row[headers["Keywords"]]
            status = row[headers["status"]]
            importance = row[headers["importance"]]
            exec_time = row[headers["exec time"]]
            steps_actions = row[headers["Steps_actions"]]
            expected_results = row[headers["expected_results"]]
            execution_type = row[headers["execution_type"]]
            test_cases_summary = row[headers["Test cases"]]



            if test_case_id:
                if current_test_case:
                    test_case_data.append(current_test_case)

                current_test_case = {
                    "Category": category,
                    "Test Case ID": test_case_id,
                    "Description": description,
                    "preconditions": preconditions,
                    "status": status,
                    "importance": importance,
                    "exec_time": exec_time,
                    "exec_type" : execution_type,
                    "Keywords": keywords,
                    "Expected Output": expected_results,
                    "Steps Data": [],
                    "summary": test_cases_summary
                }

                # Reset step number for new test case
                current_step_number = 1

            # Collect the steps for this test case
            if steps_actions:
                steps = [re.sub(r'^\d+\.\s*', '', step.strip()) for step in steps_actions.split("\n") if step.strip()]

                for action in steps:
                    step = {
                        "step_number": current_step_number,
                        "actions": action,
                        "expected_results": expected_results,
                        "execution_type": execution_type
                    }
                    current_test_case["Steps Data"].append(step)
                    current_step_number += 1


        if current_test_case:
            test_case_data.append(current_test_case)

        # Print and upload test cases
        for test_case in test_case_data:
            print(f"Category: {test_case['Category']}")
            print(f"Test Case ID: {test_case['Test Case ID']}")
            print(f"Description: {test_case['Description']}")
            print(f"preconditions: {test_case['preconditions']}")
            print(f"Keywords: {test_case['Keywords']}")
            print(f"status: {test_case['status']}")
            print(f"importance: {test_case['importance']}")
            print(f"execution type: {test_case['exec_type']}")
            print(f"exec_time: {test_case['exec_time']}")
            print(f"Expected Output: {test_case['Expected Output']}")
            print("Steps Data:", test_case['Steps Data'])
            print("Summary:", test_case['summary'])
            print("-" * 40,"\n")

            # Upload to TestLink
            self.upload_test_case_to_testlink(test_case)
        return test_case_data

    def upload_test_case_to_testlink(self,test_case):
        suite_id = self.get_or_create_test_suite(test_case['Category'])
        project_id = self.get_project_id(inputs.PROJECT_NAME)
        executiontype = test_case['exec_type']
        case_name = f"{test_case['Test Case ID']}"
        summary = f"{test_case['Description']}: {test_case['summary']}"
        preconditions = test_case['preconditions']
        status = test_case['status']
        imp = test_case['importance']
        time = test_case['exec_time']
        expected_results = test_case['Expected Output']
        steps_list = test_case['Steps Data']
        keywords_list = test_case['Keywords'].split(",") if test_case['Keywords'] else []

        # Check if the test case already exists
        if self.test_case_exists(case_name, suite_id):
            print(f"Test case '{case_name}' already exists in suite '{test_case['Category']}'. Skipping creation.")
            return

        try:
            test_case_response = self.tlc.createTestCase(
                testcasename=case_name,
                testsuiteid=suite_id,
                testprojectid=project_id,
                authorlogin="admin",
                summary=summary,
                preconditions=preconditions,
                status=status,
                importance=imp,
                estimatedexecduration=time,
                executiontype=executiontype,
                steps=steps_list,
                expectedresults=expected_results
            )
            print(f"Test case '{case_name}' created successfully.\n")

            test_case_id = test_case_response[0]['id']
            details_of_testcase = self.tlc.getTestCaseIDByName(case_name)
            test_id = details_of_testcase[0]["id"]

            tc_full_ext_id = self.tlc.getTestCase(testcaseid=test_id)[0]["full_tc_external_id"]


            if keywords_list:
                if isinstance(keywords_list, str):
                    keywords = [keyword.strip() for keyword in keywords_list.split(",") if keyword.strip()]
                else:
                    keywords = [keyword.strip() for keyword in keywords_list if keyword.strip()]
            else:
                keywords = []

            #Adding the keywords to the TestLink API
            response_keyw = self.tlc.addTestCaseKeywords({tc_full_ext_id: keywords})
            print("Keywords added:", response_keyw)


        except Exception as e:
            print(f"Error creating test case '{case_name}': {str(e)}")


    def test_case_exists(self,test_case_name, suite_id):
        # Check if the test case exists in the specified test suite
        test_cases = self.tlc.getTestCasesForTestSuite(suite_id, False, False)
        for test_case in test_cases:
            existing_name = test_case["name"].strip().lower()
            if existing_name == test_case_name.strip().lower():
                return True
        return False


    def get_project_id(self,project_name):
        projects = self.tlc.getProjects()
        for project in projects:
            if project["name"] == project_name:
                return project["id"]
        raise ValueError(f"Project '{project_name}' not found.")


    def get_or_create_test_suite(self,suite_name):
        project_id = self.get_project_id("PCI")
        suites = self.tlc.getFirstLevelTestSuitesForTestProject(project_id)
        for suite in suites:
            if suite["name"] == suite_name:
                return suite["id"]

        suite = self.tlc.createTestSuite(project_id, suite_name, "Imported test suite")
        return suite[0]["id"]

if __name__ == '__main__':
    ts = Test()
    ts.read_test_case_excel(inputs.EXCEL_PATH)
