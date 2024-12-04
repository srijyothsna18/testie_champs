import testlink
from openpyxl import load_workbook
from Core.Creations.create_test_pro import Create
from Utils.Inputs.Common_inputs import inputs

class Test:

    tc_full_ext_id = 0

    def get_project_id(self,project_name):
        projects = Create.tlc.getProjects()
        for project in projects:
            if project["name"] == project_name:
                return project["id"]
        raise ValueError(f"Project '{project_name}' not found.")


    def get_or_create_test_suite(self,project_id, suite_name):
        try:
            suites = Create.tlc.getFirstLevelTestSuitesForTestProject(project_id)
            if suites is None or len(suites) == 0:
                suite = Create.tlc.createTestSuite(project_id, suite_name, "Imported test suite")
                return suite[0]["id"]
            for suite in suites:
                if suite["name"] == suite_name:
                    return suite["id"]
            suite = Create.tlc.createTestSuite(project_id, suite_name, "Imported test suite")
            return suite[0]["id"]

        except testlink.testlinkerrors.TLResponseError as e:
            print(f"Error fetching or creating test suite: {str(e)}")
            raise

    def test_case_exists(self,test_case_name, suite_id):
        test_case_name = test_case_name.strip().lower()
        test_cases = Create.tlc.getTestCasesForTestSuite(suite_id, False, False)
        for test_case in test_cases:
            if test_case["name"].strip().lower() == test_case_name:
                return True
        return False

    def import_test_cases_from_excel(self,file_path, project_name):
        wb = load_workbook(filename=file_path)
        sheet = wb.active

        project_id = self.get_project_id(project_name)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < 6:
                continue

            category, test_id, description, keywords, steps, test_case, expected_output = row[:7]

            if not all([test_id, description, test_case, expected_output]):
                continue

            # Create or get suite ID for the category
            suite_name = category.strip()
            suite_id = self.get_or_create_test_suite(project_id, suite_name)

            # Define test case details
            case_name = f"{test_id} - {description}"
            summary = test_case.strip()
            expected_results = expected_output.strip()

            steps_list = eval(steps)  # Converts the string representation of list into actual list of dicts

            if self.test_case_exists(case_name, suite_id):
                print(f"Test case '{case_name}' already exists in suite '{suite_name}'. Skipping creation.")
                continue

            try:
                test_case_response = Create.tlc.createTestCase(
                    testcasename=case_name,
                    testsuiteid=suite_id,
                    testprojectid=project_id,
                    authorlogin="admin",
                    summary=summary,
                    steps=steps_list,
                    expectedresults=expected_results
                )
                print(f"Test case '{case_name}' created successfully in suite '{suite_name}'.","\n")

                details_of_testcase = Create.tlc.getTestCaseIDByName(case_name)
                test_id = details_of_testcase[0]["id"]
                print("test case id is ", test_id)

                self.tc_full_ext_id = Create.tlc.getTestCase(testcaseid=test_id)[0]["full_tc_external_id"]
                print("Test Case '%s' - id: %s - ext-id %s" % ("external ids", test_id, self.tc_full_ext_id))

                #Note : before adding keywords to test cases make sure you have already created keyowrd in the required projects
                if keywords:
                    if "," in keywords:
                        keywords_list = [keyword.strip() for keyword in keywords.split(",") if keyword.strip()]
                    else:
                        keywords_list = [keywords.strip()]



                response_keyw = Create.tlc.addTestCaseKeywords({self.tc_full_ext_id: keywords_list})
                print("Keywords added:", response_keyw)

            except Exception as e:
                print(f"Error creating test case '{case_name}': {str(e)}")


    # def get_requirements(self):
    #     req_list = Create.tlc.getRequirements(self.get_project_id(inputs.PROJECT_NAME))
    #     print("Existing requirements: ", req_list)
    #     reqA = req_list[0]
    #     print("requirement details: ", reqA)


    def get_requirements(self):
        # Fetch associated requirements
        req_coverage = Create.tlc.getRequirements(self.get_project_id(inputs.PROJECT_NAME))
        print(req_coverage)
        for req in req_coverage:
            print(f"Requirement: {req['srs_id']}, Test Case ID: {req['tc_id']}")

    def count_ts(self):
        result = Create.tlc.countTestCasesTP()
        print(result)
        return result

    def assign_req(self):
        for i in range(1,self.count_ts()):  # Loop from 1 to 34
            test_case = f"pcie-{i}"  # Format the test case dynamically
            response_req = Create.tlc.assignRequirements(
                test_case,
                self.get_project_id(inputs.PROJECT_NAME),
                [{'req_spec': 5359, 'requirements': [5496]}]
            )
            print(f"Assigned requirements to {test_case}")

    # def assign_req(self):
    #     response_req = Create.tlc.assignRequirements("pcie-5",self.get_project_id(inputs.PROJECT_NAME),
    #                                      [{'req_spec' : 5359,'requirements': [5496]}]
    #     )
    #     print("assignrequiremnets reqA to tc-01")


if __name__ == "__main__":
    test_instance = Test()
    test_instance.import_test_cases_from_excel(inputs.EXCEL_PATH, inputs.PROJECT_NAME)
    test_instance.get_requirements()
    test_instance.assign_req()

