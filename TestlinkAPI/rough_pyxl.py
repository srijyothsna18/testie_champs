from openpyxl import load_workbook
from testlink import TestlinkAPIClient

API_URL = "http://172.17.17.8:8085/lib/api/xmlrpc/v1/xmlrpc.php"
KEY = "10b2132073a17c9d4a0bc700dd778f83"
tlc = TestlinkAPIClient(API_URL, KEY)


def read_test_case_excel(file_path):
    # Load the Excel workbook
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active  # Use the active sheet

    # Define the columns based on headers
    headers = {cell.value: idx for idx, cell in enumerate(sheet[1])}

    # Ensure necessary columns exist
    required_columns = [
        "Category", "Test case ID", "Description", "Keywords",
        "Steps_actions", "expected_results", "execution_type"
    ]

    for column in required_columns:
        if column not in headers:
            raise ValueError(f"Missing required column: {column}")

    test_case_data = []
    current_test_case = {}

    # Iterate through rows starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):
        category = row[headers["Category"]]
        test_case_id = row[headers["Test case ID"]]
        description = row[headers["Description"]]
        keywords = row[headers["Keywords"]]
        steps_actions = row[headers["Steps_actions"]]
        expected_results = row[headers["expected_results"]]
        execution_type = row[headers["execution_type"]]

        # Check if we are starting a new test case
        if test_case_id:
            # If we have a current test case, append it to the list
            if current_test_case:
                test_case_data.append(current_test_case)

            # Start a new test case dictionary
            current_test_case = {
                "Category": category,
                "Test Case ID": test_case_id,
                "Description": description,
                "Keywords": keywords,
                "Test Cases": description,  # Assuming description doubles as test cases
                "Expected Output": expected_results,
                "Steps Data": []  # Initialize Steps Data as an empty list
            }

        # Collect the steps for this test case
        if steps_actions:
            step = {
                "step_number": len(current_test_case["Steps Data"]) + 1,
                "actions": steps_actions,
                "expected_results": expected_results,
                "execution_type": execution_type
            }
            current_test_case["Steps Data"].append(step)

    # Append the last test case after the loop ends
    if current_test_case:
        test_case_data.append(current_test_case)

    # Print and upload test cases
    for test_case in test_case_data:
        print(f"Category: {test_case['Category']}")
        print(f"Test Case ID: {test_case['Test Case ID']}")
        print(f"Description: {test_case['Description']}")
        print(f"Keywords: {test_case['Keywords']}")
        print(f"Test Cases: {test_case['Test Cases']}")
        print(f"Expected Output: {test_case['Expected Output']}")
        print("Steps Data:", test_case['Steps Data'])
        print("-" * 40)

        # Upload to TestLink
        upload_test_case_to_testlink(test_case)


def upload_test_case_to_testlink(test_case):
    suite_id = get_or_create_test_suite(test_case['Category'])
    project_id = get_project_id('Your Project Name')  # Replace with actual project name

    case_name = f"{test_case['Test Case ID']} - {test_case['Description']}"
    summary = test_case['Test Cases']
    expected_results = test_case['Expected Output']
    steps_list = test_case['Steps Data']
    keywords_list = test_case['Keywords'].split(",") if test_case['Keywords'] else []

    try:
        test_case_response = tlc.createTestCase(
            testcasename=case_name,
            testsuiteid=suite_id,
            testprojectid=project_id,
            authorlogin="admin",  # Change if necessary
            summary=summary,
            steps=steps_list,
            expectedresults=expected_results
        )
        print(f"Test case '{case_name}' created successfully.")

        test_case_id = test_case_response[0]['id']
        print("Test case ID:", test_case_id)

        if keywords_list:
            response_keywords = tlc.addTestCaseKeywords({test_case_id: keywords_list})
            print("Keywords added:", response_keywords)

    except Exception as e:
        print(f"Error creating test case '{case_name}': {str(e)}")


def get_project_id(project_name):
    projects = tlc.getProjects()
    for project in projects:
        if project["name"] == "PCI":
            return project["id"]
    raise ValueError(f"Project '{project_name}' not found.")


def get_or_create_test_suite(suite_name):
    project_id = get_project_id("PCI")  # Replace with your project name
    suites = tlc.getFirstLevelTestSuitesForTestProject(project_id)
    for suite in suites:
        if suite["name"] == suite_name:
            return suite["id"]

    suite = tlc.createTestSuite(project_id, suite_name, "Imported test suite")
    return suite[0]["id"]


# File path to your Excel sheet
file_path = 'TC_Excel_sheet/Copy_of_NVMe_Test_cases.xlsx'
read_test_case_excel(file_path)
